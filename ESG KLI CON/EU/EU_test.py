import urllib.parse
from bs4 import BeautifulSoup
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from datetime import datetime
from openpyxl import load_workbook
import pandas as pd
import time
import re
import urllib.parse as urlparse
import os

error_list = []
seen_titles = set()
seen_urls = set()
forbidden_keywords = [
    "amendment", "amending", "to amend",
    "modifying", "to modify","appropriation",
    "budget","repealed","revoked","airport","airline",
    "appointment", "appointed","patient","coronavirus", "COVID-19"
]

def getSoup(url):
    url_re = requests.get(url)
    url_con = BeautifulSoup(url_re.content,'html.parser')
    return url_con

def get_total_pages_from_form(html):
    base_url = "https://eur-lex.europa.eu/"

    form = html.find("form", {"id": "pagingFormtop"})
    if not form:
        return 1 

    input_tag = form.find("input", {"id": "pagingInput1"})
    if input_tag and "onkeyup" in input_tag.attrs:
        match = re.search(r"checkPagingFO\([^,]+,\s*(\d+)\)", input_tag["onkeyup"])
        if match:
            return int(match.group(1))

    page_numbers = []
    for a in form.find_all("a", href=True):
        match = re.search(r"page=(\d+)", a["href"])
        if match:
            page_numbers.append(int(match.group(1)))

    if page_numbers:
        return max(page_numbers)
    return 1

def url_generator(counter, base_url):
    if counter == 1:
        return link_list
    else:
        counter_start = 2
        for page in range(counter_start, counter + 1):
            paginated_url = f"{base_url}&page={page}"
            link_list.append(paginated_url)
        return link_list

def article_continer_page_extraction(soup_of_article_container_page):
    functioned_article_list = []
    try:
        find_article_container = soup_of_article_container_page.find("div", id="EurlexContent")
        if find_article_container:
            all_articles_of_container = find_article_container.find_all("div", class_="SearchResult")
            if all_articles_of_container:
                for article_ in all_articles_of_container:
                    functioned_article_list.append(article_)
                return functioned_article_list
            else:
                return functioned_article_list
        else:
            return functioned_article_list
    except Exception as e:
        print(f"Error in article_continer_page_extraction: {e}")
        return functioned_article_list

def check_wheather_empty(main_functioned_articles):
    try:
        if not main_functioned_articles:
            print(f"👎  Total articles collected: {len(main_functioned_articles)}")
            print("There is no articles to proceed 🛑")
            return False
        print(f"✌️  Total articles collected: {len(main_functioned_articles)}")
        return True 
    except Exception as e:
        print(f"Error in check_wheather_empty: {e}")
        return False

def extract_basic_article_link(html: str, base_url: str = "https://eur-lex.europa.eu/") -> str | None:
    a_tag = html.find("a", class_="title")
    if a_tag and a_tag.get("href"):
        return urljoin(base_url, a_tag["href"])
    return None

def get_document_info_link(html: str, base_url: str = "https://eur-lex.europa.eu") -> str | None:
    sidebar = html.find("div", class_="col-md-3 sidebar-offcanvas")

    if sidebar:
        a_tag = sidebar.find("a", string=lambda t: t and "Document information" in t)
        if a_tag:
            return urljoin(base_url, a_tag.get("href"))
    return None

def convert_date_format(date_str: str) -> str:
    try:
        return datetime.strptime(date_str, "%d/%m/%Y").strftime("%Y-%m-%d")
    except ValueError:
        return date_str

def data_scraping_of_an_article(soup_of_details_page,document_details_page_url_):
    article_details_ = []
    title = type_of_regulation = date_of_document = date_of_effect = None
    source_url = document_details_page_url_
    find_main_content_container = soup_of_details_page.find("div", class_="EurlexContent")
    if find_main_content_container:
        title_ = find_main_content_container.find("p", id="title")
        if title_:
            title = title_.get_text(strip = True)

        p_tags = find_main_content_container.find_all("p")
        for p in p_tags:
            if p.get_text(strip=True).startswith("ELI:"):
                a_tag = p.find("a", href=True)
                if a_tag:
                    source_url = a_tag["href"]

        multilingual_point = find_main_content_container.find('div', id='multilingualPoint')
        panels = multilingual_point.find_all('div', class_='panel panel-default PagePanel')
        misc_panel = None
        for panel in panels:
            heading = panel.find('div', class_='panel-heading')
            if heading and 'Miscellaneous information' in heading.get_text():
                misc_panel = panel
                break
        if misc_panel:
            misc_contents = misc_panel.find('div', id='PPMisc_Contents')
            if misc_contents:
                form_dt = misc_contents.find('dt', string=lambda text: text and 'Form:' in text)
                if form_dt:
                    form_dd = form_dt.find_next('dd')
                    if form_dd:
                        span = form_dd.find('span')
                        if span:
                            type_of_regulation = span.get_text(strip=True)
                        else:
                            type_of_regulation = form_dd.get_text(strip=True)
                    else:
                        print("Form dd element not found")
                else:
                    print("Form dt element not found")
            else:
                print("PPMisc_Contents not found")
        else:
            print("Miscellaneous information panel not found")

        dates_panel = None
        for panel in panels:
            heading = panel.find('div', class_='panel-heading')
            if heading and 'Dates' in heading.get_text():
                dates_panel = panel
                break

        if dates_panel:
            dates_contents = dates_panel.find('div', id='PPDates_Contents')
            
            if dates_contents:
                date_dt = dates_contents.find('dt', string=lambda text: text and 'Date of document:' in text)
                
                if date_dt:
                    date_dd = date_dt.find_next('dd')
                    
                    if date_dd:
                        date_of_document = convert_date_format(date_dd.get_text(strip=True).split(';')[0])
                    else:
                        print("Date dd element not found")
                else:
                    print("Date of document dt element not found")

                date_doe = dates_contents.find('dt', string=lambda text: text and 'Date of effect:' in text)
                
                if date_doe:
                    date_de = date_doe.find_next('dd')
                    
                    if date_de:
                        date_of_effect = convert_date_format(date_de.get_text(strip=True).split(';')[0])
                    else:
                        print("Date dd element not found")
                else:
                    print("Date of effect dt element not found")
            else:
                print("PPDates_Contents not found")
        else:
            print("Dates panel not found")
        
        print(f"\nOriginal Title       : {title}")
        print(f"Source Link          : {source_url}")
        print(f"Type of Regulation   : {type_of_regulation}") 
        print(f"Date of Adoption     : {date_of_document}")
        print(f"Entry into Force Date: {date_of_effect}")

        if title in seen_titles or source_url in seen_urls:
            print(f"♻️  Duplicate skipped → Title: {title}, URL: {source_url}")
            return None 

        if title:
            lower_title = title.lower()
            if any(kw in lower_title for kw in forbidden_keywords):
                print(f"⏭️  Skipped due to forbidden keyword in title → {title}")
                return None
            
        if title:
            seen_titles.add(title)
        if source_url:
            seen_urls.add(source_url)

        article_details_.append({
            "Original Title": title,
            "Source Link": source_url,
            "Type of Regulation": type_of_regulation,
            "Date of Adoption": date_of_document,
            "Entry into Force Date": date_of_effect
        })

        return article_details_ 

def sum_numbers(new_value=0, _storage={"total": 0}):
    _storage["total"] += new_value
    return _storage["total"]

def custom_text_to_hex(text):
    try:
        text_with_plus = text.replace(" ", "+")
        encoded = urllib.parse.quote(text_with_plus, safe='+')
        return f"%22{encoded}%22"
    except Exception as e:
        msg = f"Encoding error for text '{text}': {e}"
        error_list.append(msg)
        return "%22%22"
# Function Part is over

output_file = "EU_related_out.xlsx"
if os.path.exists(output_file):
    try:
        os.remove(output_file)
        print(f"🗑️  Existing file '{output_file}' removed to start fresh.")
    except Exception as e:
        print(f"⚠️  Could not remove existing file '{output_file}': {e}")

try:
    with open('key_words.txt', 'r', encoding='utf-8') as file:
        keyword_list = file.read().strip().splitlines()
except Exception as error:
    msg = f"Error reading keyword file: {error}"
    error_list.append(msg)
    keyword_list = []

if error_list:
    print("Errors:", error_list)

if not keyword_list:
    print("No keywords found.")
else:
    for sin_key in keyword_list:
        key_related_articles = []
        link_list = []
        print(f"\n\n🌟 {sin_key}'s Details 🌟")
        list_of_links_each_key = []
        list_of_articles = []
        main_functioned_articles = []

        try:
            normal_input = f"{sin_key}"
            hex_output = custom_text_to_hex(normal_input)

            url_for_qid = f"https://eur-lex.europa.eu/search.html?VV=true&SUBDOM_INIT=LEGISLATION&DTS_SUBDOM=LEGISLATION&textScope1=ti&DTS_DOM=EU_LAW&lang=en&type=advanced&andText1={hex_output}"
            response = requests.get(url_for_qid, allow_redirects=True, timeout=20)

            final_url = response.url
            qid = urlparse.parse_qs(urlparse.urlparse(final_url).query).get("qid", [None])[0]

            if not qid:
                print(f"⚠️ No qid found for {sin_key}, skipping...")
                continue

            base_url_common = f"https://eur-lex.europa.eu/search.html?VV=true&SUBDOM_INIT=LEGISLATION&DTS_SUBDOM=LEGISLATION&textScope1=ti&DTS_DOM=EU_LAW&lang=en&type=advanced&qid={qid}&wh0=andCOMPOSE%3DENG%2CorEMBEDDED_MANIFESTATION-TYPE%3Dpdf%3BEMBEDDED_MANIFESTATION-TYPE%3Dpdfa1a%3BEMBEDDED_MANIFESTATION-TYPE%3Dpdfa1b%3BEMBEDDED_MANIFESTATION-TYPE%3Dpdfa2a%3BEMBEDDED_MANIFESTATION-TYPE%3Dpdfx%3BEMBEDDED_MANIFESTATION-TYPE%3Dpdf1x%3BEMBEDDED_MANIFESTATION-TYPE%3Dhtml%3BEMBEDDED_MANIFESTATION-TYPE%3Dxhtml%3BEMBEDDED_MANIFESTATION-TYPE%3Ddoc%3BEMBEDDED_MANIFESTATION-TYPE%3Ddocx&andText1={hex_output}"

            if base_url_common:
                try:
                    first_page_content = getSoup(base_url_common)
                    count = get_total_pages_from_form(first_page_content)
                    link_list.append(base_url_common)
                    related_urls = url_generator(count, base_url_common)

                    for link in related_urls:
                        try:
                            soup_of_each_link = getSoup(link)
                            returned_articles = article_continer_page_extraction(soup_of_each_link)
                            if returned_articles:
                                main_functioned_articles.extend(returned_articles)
                        except Exception as e:
                            msg = f"Error processing related URL {link}: {e}"
                            print(msg)
                            error_list.append(msg)

                    articles_contained_keywords = check_wheather_empty(main_functioned_articles)
                    if articles_contained_keywords:
                        print("Proceeding with processing...🚀")
                        print(f"🌐  Filters and Keyword added URL:\n{base_url_common}")
                        for article_results in main_functioned_articles:
                            try:
                                basic_article_link = extract_basic_article_link(article_results)
                                each_article_base_page_soup = getSoup(basic_article_link)
                                document_details_page_url = get_document_info_link(each_article_base_page_soup)
                                article_document_page_soup = getSoup(document_details_page_url)
                                article_data_ = data_scraping_of_an_article(article_document_page_soup,document_details_page_url)
                                if article_data_:
                                    key_related_articles.extend(article_data_)
                            except Exception as e:
                                msg = f"Error extracting article details: {e}"
                                print(msg)
                                error_list.append(msg)
                    else:
                        print(f"🌐  Filters and Keyword added URL:\n{base_url_common}")
                        pass
                except Exception as e:
                    msg = f"Error in main scraping loop for {sin_key}: {e}"
                    print(msg)
                    error_list.append(msg)

        except Exception as e:
            msg = f"Error preparing URL for {sin_key}: {e}"
            print(msg)
            error_list.append(msg)

        try:
            if key_related_articles:
                df = pd.DataFrame(key_related_articles)
                try:
                    book = load_workbook("EU_related_out.xlsx")
                    with pd.ExcelWriter("EU_related_out.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                        start_row = writer.sheets["Sheet1"].max_row
                        df.to_excel(writer, index=False, header=False, startrow=start_row)
                except FileNotFoundError:
                    df.to_excel("EU_related_out.xlsx", index=False, engine="openpyxl")

                print(f"\n✅  {sin_key}'s {len(key_related_articles)} Results appended to 'EU_related_out.xlsx'")
                sum_numbers(len(key_related_articles))
            else:
                print(f"⚠️  {sin_key}'s (Count: {len(key_related_articles)}). No results to save.")
        except Exception as e:
            msg = f"Error saving Excel file for {sin_key}: {e}"
            print(msg)
            error_list.append(msg)

try:
    sum_ = sum_numbers()
    print(f"\n📂  {sum_} rows appended to Excel.")
except TypeError as e:
    print(f"⚠️ No value passed to sum_numbers: {e}")
    sum_ = 0
    print(f"\n📂❌  {sum_} rows appended to Excel.")
